import openpyxl
import xlsxwriter
import pandas as pd
from numpy import log
from math import e
from random import random
from scipy.stats import chi2, poisson, norm

mult = 16843009
adi = 826366247
m = (2**32)-1
x0 = 1357


def getPositiveOrZeroInt(mensaje):
    while True:
        var = input(mensaje)
        try:
            var = int(var)
            if var >= 0:
                return var
        except:
            print("Valor fuera del rango valido")


def getPositiveInt(mensaje):
    while True:
        var = input(mensaje)
        try:
            var = int(var)
            if var > 0:
                return var
        except:
            print("Valor fuera del rango valido")


def getPositiveFloat(mensaje):
    while True:
        var = input(mensaje)
        try:
            var = float(var)
            if var > 0:
                return var
        except:
            print("Valor fuera del rango valido")


# Genera variables Poisson contando cuantas veces tuve que multiplicar 1 por un numero aleatorio uniformemente
# distribuido entre [0;1) para que fuera menor que e^lambda
def getVarPoisson(lamb):
    asdf = e**-lamb
    numero = -1
    b = 1
    while b > asdf:
        b = b*random()
        numero += 1
    return numero


# Dada la media y la varianza calcula una variable aleatoria Normal restandole 6 a la sumatoria de 12 numeros
# aleatorios uniformemente distribuidos entre [0;1) y eso lo multriplica por la desviacion estandar, y le suma
# la media
def getVarNormal(med, var):
    randoms = [random() for i in range(12)]
    suma = sum(randoms)
    num = ((suma-6)*(var**0.5)) + med
    return num


# Dados los valores minimos y maximos que se esperan de la muestra, se calcula una variable aleatoria normal
# multiplicando la diferencia entre el maximo y el minimo por un numero aleatorio uniformemente distribuido entre [0;1),
# y le suma el minimo
def getVarUniforme(a, b):
    num = (b-a) * random() + a
    return num


# Cociente de 1 negativo y la media o la varianza de la muestra, multiplicado por el logaritmo natural de un numero
# aleatorio uniformemente distribuido entre [0;1)
def getVarExponencial(lamb):
    num = -1/lamb*log(random())
    return num


def generateExponenciales(cant, intervalos, var_lambda):
    workbook = xlsxwriter.Workbook("tp2.xlsx", {'constant_memory': True})
    worksheet = workbook.add_worksheet("data")
    col_headers = ["Variable"]
    for i in range(len(col_headers)):
        worksheet.write(0, i, col_headers[i])
    l = var_lambda
    for i in range(1, cant+1):
        for j in range(len(col_headers)):
            num = getVarExponencial(l)
            worksheet.write(i, j, num)
    workbook.close()
    df1 = pd.read_excel("tp2.xlsx", sheet_name="data", header=0, usecols="A")
    maxVal = df1["Variable"].max()
    minVal = df1["Variable"].min()
    split = maxVal - minVal
    paso = split / intervalos
    limInf = minVal
    limSup = minVal + paso
    diccionario = []
    for i in range(intervalos):
        # prob Acumulada del lim superior menos la prob acumulada del inferior, multiplicada por la cantidad generada
        # para tener la cantidad esperada y no la probabilidad
        frecEsperadaSup = 1-e**(-limSup*l)
        frecEsperadaInf = 1-e**(-limInf*l)
        frecEsperada = (frecEsperadaSup-frecEsperadaInf)*cant
        promedio = round((limInf + limSup) / 2, 4)
        diccionario.append({"Lim Inferior": limInf,
                            "Lim Superior": limSup,
                            "Marca de Clase": promedio,
                            "Frecuencia Observada": 0,
                            "Frecuencia Esperada": frecEsperada,
                            })
        # A la base le agrego el paso para el proximo intervalo a definir
        limSup += paso
        limInf += paso
    # Cuento las frecuencias de las variables aleatorias
    for i in range(df1.shape[0]):
        for fila in diccionario:
            if fila["Lim Superior"] > df1.iloc[i, 0] >= fila["Lim Inferior"]:
                fila["Frecuencia Observada"] += 1
    chiacumulado = 0
    for fila in diccionario:
        chiacumulado += ((fila["Frecuencia Observada"]-fila["Frecuencia Esperada"])**2)/fila["Frecuencia Esperada"]
    probnook = round(chi2.cdf(chiacumulado, intervalos - 1), 4)
    writer = pd.ExcelWriter("tp2.xlsx", engine="openpyxl", mode="a")
    df2 = pd.DataFrame(diccionario)
    df2.to_excel(writer, sheet_name="estadistica")
    df4 = pd.DataFrame(columns=["Chi Calculado", "Grados de libertad", "1-P", "P"])
    df4 = pd.concat([df4, pd.DataFrame([{"Chi Calculado": chiacumulado, "Grados de libertad": intervalos - 1, "1-P": probnook,
                                         "P": round(1 - probnook, 4)}])], ignore_index=True)
    df4.to_excel(writer, sheet_name="Chi Cuadrado")
    writer.close()

    tp2excel = openpyxl.open("tp2.xlsx")
    estadistica_sheet = tp2excel["estadistica"]
    chart = openpyxl.chart.BarChart()
    chart.type = "col"
    chart.style = 10
    chart.y_axis.title = "Frecuencia"
    chart.x_axis.title = "Promedio Intervalo"
    # Promedios
    cats = openpyxl.chart.Reference(estadistica_sheet, min_col=4, min_row=2, max_row=intervalos+1)
    # Frecuencias
    data = openpyxl.chart.Reference(estadistica_sheet, min_col=5, max_col=5, min_row=1, max_row=intervalos+1)
    data2 = openpyxl.chart.Reference(estadistica_sheet, min_col=6, max_col=6, min_row=1, max_row=intervalos+1)
    chart.add_data(data, titles_from_data=True)
    chart.add_data(data2, titles_from_data=True)
    chart.set_categories(cats)
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.shape = 4
    estadistica_sheet.add_chart(chart, "K1")
    tp2excel.save("tp2.xlsx")

def generatePoissones(cant, intervalos, var_lambda):
    workbook = xlsxwriter.Workbook("tp2.xlsx", {'constant_memory': True})
    worksheet = workbook.add_worksheet("data")
    col_headers = ["Variable"]
    for i in range(len(col_headers)):
        worksheet.write(0, i, col_headers[i])
    l = var_lambda
    for i in range(1, cant+1):
        for j in range(len(col_headers)):
            num = getVarPoisson(l)
            worksheet.write(i, j, num)
    workbook.close()
    df1 = pd.read_excel("tp2.xlsx", sheet_name="data", header=0, usecols="A")
    maxVal = df1["Variable"].max()
    minVal = df1["Variable"].min()
    split = maxVal - minVal
    paso = split / intervalos
    limInf = minVal
    limSup = minVal + paso
    diccionario = []
    for i in range(intervalos):
        # prob Acumulada del lim superior menos la prob acumulada del inferior, multiplicada por la cantidad generada
        # para tener la cantidad esperada y no la probabilidad
        frecEsperadaSup = poisson.cdf(limSup, l)
        frecEsperadaInf = poisson.cdf(limInf, l)
        frecEsperada = round((frecEsperadaSup-frecEsperadaInf)*cant, 4)
        promedio = round((limInf + limSup) / 2, 4)
        diccionario.append({"Lim Inferior": limInf,
                            "Lim Superior": limSup,
                            "Marca de Clase": promedio,
                            "Frecuencia Observada": 0,
                            "Frecuencia Esperada": frecEsperada,
                            })
        # A la base le agrego el paso para el proximo intervalo a definir
        limSup += paso
        limInf += paso
    # Cuento las frecuencias de las variables aleatorias
    for i in range(df1.shape[0]):
        for fila in diccionario:
            if fila == diccionario[intervalos-1]:
                if fila["Lim Superior"] >= df1.iloc[i, 0] >= fila["Lim Inferior"]:
                    fila["Frecuencia Observada"] += 1
            else:
                if fila["Lim Superior"] > df1.iloc[i, 0] >= fila["Lim Inferior"]:
                    fila["Frecuencia Observada"] += 1
    chiacumulado = 0
    for fila in diccionario:
        chiacumulado += ((fila["Frecuencia Observada"]-fila["Frecuencia Esperada"])**2)/fila["Frecuencia Esperada"]
    probnook = round(chi2.cdf(chiacumulado, intervalos - 1), 4)
    writer = pd.ExcelWriter("tp2.xlsx", engine="openpyxl", mode="a")
    df2 = pd.DataFrame(diccionario)
    df2.to_excel(writer, sheet_name="estadistica")
    df4 = pd.DataFrame(columns=["Chi Calculado", "Grados de libertad", "1-P", "P"])
    df4 = pd.concat([df4, pd.DataFrame([{"Chi Calculado": chiacumulado, "Grados de libertad": intervalos - 1, "1-P": probnook,
                                         "P": round(1 - probnook, 4)}])], ignore_index=True)
    df4.to_excel(writer, sheet_name="Chi Cuadrado")
    writer.close()
    tp2excel = openpyxl.open("tp2.xlsx")
    estadistica_sheet = tp2excel["estadistica"]
    chart = openpyxl.chart.BarChart()
    chart.type = "col"
    chart.style = 10
    chart.y_axis.title = "Frecuencia"
    chart.x_axis.title = "Promedio Intervalo"
    # Promedios
    cats = openpyxl.chart.Reference(estadistica_sheet, min_col=4, min_row=2, max_row=intervalos+1)
    # Frecuencias
    data = openpyxl.chart.Reference(estadistica_sheet, min_col=5, max_col=5, min_row=1, max_row=intervalos+1)
    data2 = openpyxl.chart.Reference(estadistica_sheet, min_col=6, max_col=6, min_row=1, max_row=intervalos+1)
    chart.add_data(data, titles_from_data=True)
    chart.add_data(data2, titles_from_data=True)
    chart.set_categories(cats)
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.shape = 4
    estadistica_sheet.add_chart(chart, "K1")
    tp2excel.save("tp2.xlsx")

def generateNormales(cant, intervalos, media, varianza):
    workbook = xlsxwriter.Workbook("tp2.xlsx", {'constant_memory': True})
    worksheet = workbook.add_worksheet("data")
    col_headers = ["Variable"]
    for i in range(len(col_headers)):
        worksheet.write(0, i, col_headers[i])
    u = media
    v = varianza
    for i in range(1, cant+1):
        for j in range(len(col_headers)):
            num = getVarNormal(u, v)
            worksheet.write(i, j, num)
    workbook.close()
    df1 = pd.read_excel("tp2.xlsx", sheet_name="data", header=0, usecols="A")
    maxVal = df1["Variable"].max()
    minVal = df1["Variable"].min()
    split = maxVal - minVal
    paso = split / intervalos
    limInf = minVal
    limSup = minVal + paso
    diccionario = []
    for i in range(intervalos):
        # prob Acumulada del lim superior menos la prob acumulada del inferior, multiplicada por la cantidad generada
        # para tener la cantidad esperada y no la probabilidad
        frecEsperadaSup = norm.cdf(limSup, u, v**0.5)
        frecEsperadaInf = norm.cdf(limInf, u, v**0.5)
        frecEsperada = round((frecEsperadaSup-frecEsperadaInf)*cant, 4)
        promedio = round((limInf + limSup) / 2, 4)
        diccionario.append({"Lim Inferior": limInf,
                            "Lim Superior": limSup,
                            "Marca de Clase": promedio,
                            "Frecuencia Observada": 0,
                            "Frecuencia Esperada": frecEsperada,
                            })
        # A la base le agrego el paso para el proximo intervalo a definir
        limSup += paso
        limInf += paso
    # Cuento las frecuencias de las variables aleatorias
    for i in range(df1.shape[0]):
        for fila in diccionario:
            if fila == diccionario[intervalos-1]:
                if fila["Lim Superior"] >= df1.iloc[i, 0] >= fila["Lim Inferior"]:
                    fila["Frecuencia Observada"] += 1
            else:
                if fila["Lim Superior"] > df1.iloc[i, 0] >= fila["Lim Inferior"]:
                    fila["Frecuencia Observada"] += 1
    chiacumulado = 0
    for fila in diccionario:
        chiacumulado += ((fila["Frecuencia Observada"]-fila["Frecuencia Esperada"])**2)/fila["Frecuencia Esperada"]
    probnook = round(chi2.cdf(chiacumulado, intervalos - 1), 4)
    writer = pd.ExcelWriter("tp2.xlsx", engine="openpyxl", mode="a")
    df2 = pd.DataFrame(diccionario)
    df2.to_excel(writer, sheet_name="estadistica")
    df4 = pd.DataFrame(columns=["Chi Calculado", "Grados de libertad", "1-P", "P"])
    df4 = pd.concat([df4, pd.DataFrame([{"Chi Calculado": chiacumulado, "Grados de libertad": intervalos - 1, "1-P": probnook,
                                         "P": round(1 - probnook, 4)}])], ignore_index=True)
    df4.to_excel(writer, sheet_name="Chi Cuadrado")
    writer.close()
    tp2excel = openpyxl.open("tp2.xlsx")
    estadistica_sheet = tp2excel["estadistica"]
    chart = openpyxl.chart.BarChart()
    chart.type = "col"
    chart.style = 10
    chart.y_axis.title = "Frecuencia"
    chart.x_axis.title = "Promedio Intervalo"
    # Promedios
    cats = openpyxl.chart.Reference(estadistica_sheet, min_col=4, min_row=2, max_row=intervalos+1)
    # Frecuencias
    data = openpyxl.chart.Reference(estadistica_sheet, min_col=5, max_col=5, min_row=1, max_row=intervalos+1)
    data2 = openpyxl.chart.Reference(estadistica_sheet, min_col=6, max_col=6, min_row=1, max_row=intervalos+1)
    chart.add_data(data, titles_from_data=True)
    chart.add_data(data2, titles_from_data=True)
    chart.set_categories(cats)
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.shape = 4
    estadistica_sheet.add_chart(chart, "K1")
    tp2excel.save("tp2.xlsx")

def generateUniformes(cant, intervalos, valor_min, valor_max):
    workbook = xlsxwriter.Workbook("tp2.xlsx", {'constant_memory': True})
    worksheet = workbook.add_worksheet("data")
    col_headers = ["Variable"]
    for i in range(len(col_headers)):
        worksheet.write(0, i, col_headers[i])
    a = valor_min
    b = valor_max
    for i in range(1, cant+1):
        for j in range(len(col_headers)):
            num = getVarUniforme(a, b)
            worksheet.write(i, j, num)
    workbook.close()
    df1 = pd.read_excel("tp2.xlsx", sheet_name="data", header=0, usecols="A")
    # Busco los valores max y min de la muestra
    maxVal = df1["Variable"].max()
    minVal = df1["Variable"].min()
    # Calculo la diferencia entre los valores max y min
    split = maxVal - minVal
    # Calculo el paso entre intervalos
    paso = split / intervalos
    # Seteo el inicio de los intervalos
    limInf = minVal
    limSup = minVal + paso
    diccionario = []
    # Armo una lista de diccionarios con la información de los intervalos
    for i in range(intervalos):
        # Como es uniforme la frecuencia esperada es el cociente entre la cantidad de variables y
        # la cantidad de intervalos
        frecEsperada = cant/intervalos
        promedio = round((limInf + limSup) / 2, 4)
        diccionario.append({"Lim Inferior": limInf,
                            "Lim Superior": limSup,
                            "Marca de Clase": promedio,
                            "Frecuencia Observada": 0,
                            "Frecuencia Esperada": frecEsperada,
                            })
        # A la base le agrego el paso para el proximo intervalo a definir
        limSup += paso
        limInf += paso
    # Cuento las frecuencias de las variables aleatorias
    for i in range(df1.shape[0]):
        for fila in diccionario:
            if fila == diccionario[intervalos-1]:
                if fila["Lim Superior"] >= df1.iloc[i, 0] >= fila["Lim Inferior"]:
                    fila["Frecuencia Observada"] += 1
            else:
                if fila["Lim Superior"] > df1.iloc[i, 0] >= fila["Lim Inferior"]:
                    fila["Frecuencia Observada"] += 1
    chiacumulado = 0
    # Calculo los Chi^2
    for fila in diccionario:
        chiacumulado += ((fila["Frecuencia Observada"]-fila["Frecuencia Esperada"])**2)/fila["Frecuencia Esperada"]
    # Calculo alfa(lo calcula scipy)
    probnook = round(chi2.cdf(chiacumulado, intervalos - 1), 4)
    writer = pd.ExcelWriter("tp2.xlsx", engine="openpyxl", mode="a")
    df2 = pd.DataFrame(diccionario)
    # Esto arma la hoja "estadistica" en el excel
    df2.to_excel(writer, sheet_name="estadistica")
    df4 = pd.DataFrame(columns=["Chi Calculado", "Grados de libertad", "1-P", "P"])
    df4 = pd.concat([df4, pd.DataFrame([{"Chi Calculado": chiacumulado, "Grados de libertad": intervalos - 1, "1-P": probnook,
                                        "P": round(1 - probnook, 4)}])], ignore_index=True)
    # Esto arma la hoja Chi Cuadrado en el excel
    df4.to_excel(writer, sheet_name="Chi Cuadrado")
    writer.close()
    tp2excel = openpyxl.open("tp2.xlsx")
    estadistica_sheet = tp2excel["estadistica"]
    # Arranco el grafico
    chart = openpyxl.chart.BarChart()
    chart.type = "col"
    chart.style = 10
    chart.y_axis.title = "Frecuencia"
    chart.x_axis.title = "Promedio Intervalo"
    # Promedios
    cats = openpyxl.chart.Reference(estadistica_sheet, min_col=4, min_row=2, max_row=intervalos+1)
    # Frecuencias
    data = openpyxl.chart.Reference(estadistica_sheet, min_col=5, max_col=5, min_row=1, max_row=intervalos+1)
    data2 = openpyxl.chart.Reference(estadistica_sheet, min_col=6, max_col=6, min_row=1, max_row=intervalos+1)
    chart.add_data(data, titles_from_data=True)
    chart.add_data(data2, titles_from_data=True)
    chart.set_categories(cats)
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.shape = 4
    estadistica_sheet.add_chart(chart, "K1")
    tp2excel.save("tp2.xlsx")





